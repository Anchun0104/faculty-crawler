from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from faculty_workflow.database import WorkflowDatabase
from faculty_workflow.models import normalize_email, normalize_key, normalize_url


FINAL_HEADERS = ("全名", "邮箱", "last name", "职称", "工作单位（学校）", "所在院系", "个人主页", "专业方向")
REVIEW_HEADERS = (
    "ID", "状态", "全名", "邮箱", "last name", "原始职称", "规范职称", "学校",
    "院系", "个人主页", "专业方向", "来源页", "复核原因", "字段证据", "审核备注",
)
COMPLETED_EVIDENCE_HEADERS = FINAL_HEADERS + (
    "Original title", "Translated title", "Title language", "Translation status",
    "Translation engine", "Classification rules version", "Evidence URLs",
    "Primary source URL",
)
REVIEW_EVIDENCE_HEADERS = REVIEW_HEADERS + (
    "Translated title", "Title language", "Translation status", "Translation engine",
    "Classification rules version", "Evidence URLs",
)


def export_task(database: WorkflowDatabase, task_id: str, output_dir: str | Path | None = None) -> dict[str, Path]:
    task = database.get_task(task_id)
    root = Path(output_dir or task["output_dir"])
    root.mkdir(parents=True, exist_ok=True)
    accepted = database.list_candidates(task_id, ["accepted"])
    review = database.list_candidates(task_id, ["review", "candidate"])
    _validate_accepted(accepted)

    final_path = root / f"{task_id}_教授信息_飞书导入版.xlsx"
    completed_path = root / f"{task_id}_completed_evidence.xlsx"
    review_path = root / f"{task_id}_review_queue.xlsx"
    audit_path = root / f"{task_id}_audit.json"
    _write_final_workbook(accepted, final_path)
    _write_completed_evidence_workbook(accepted, completed_path)
    _write_review_workbook(review, review_path)
    audit = _build_audit(database, task_id, accepted, review)
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    _verify_final_workbook(final_path, len(accepted))
    return {"final": final_path, "completed": completed_path, "review": review_path, "audit": audit_path}


def _write_final_workbook(rows: Iterable[Any], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "学者信息"
    _append_text_row(sheet, FINAL_HEADERS)
    for row in rows:
        _append_text_row(
            sheet,
            (
                row["name"], row["email"], row["last_name"], row["normalized_title"],
                row["school"], row["department"], row["homepage"], row["direction"],
            ),
        )
        cell = sheet.cell(row=sheet.max_row, column=7)
        cell.hyperlink = row["homepage"]
        cell.style = "Hyperlink"
    sheet.freeze_panes = "A2"
    workbook.save(path)


def _write_review_workbook(rows: Iterable[Any], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "待复核"
    _append_text_row(sheet, REVIEW_EVIDENCE_HEADERS)
    for row in rows:
        _append_text_row(
            sheet,
            (
                str(row["id"]), row["status"], row["name"], row["email"], row["last_name"],
                row["title_raw"], row["normalized_title"], row["school"], row["department"],
                row["homepage"], row["direction"], row["source_url"], row["review_reason"],
                row["evidence_json"], row["decision_note"],
                row["title_translated"], row["title_language"], row["translation_status"],
                row["translation_engine"], row["classification_rules_version"],
                _evidence_urls(row["evidence_json"]),
            ),
        )
        for column in (10, 12):
            cell = sheet.cell(row=sheet.max_row, column=column)
            if normalize_url(str(cell.value or "")):
                cell.hyperlink = str(cell.value)
                cell.font = Font(color="0563C1", underline="single")
    workbook.save(path)


def _write_completed_evidence_workbook(rows: Iterable[Any], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Completed evidence"
    _append_text_row(sheet, COMPLETED_EVIDENCE_HEADERS)
    for row in rows:
        _append_text_row(
            sheet,
            (
                row["name"], row["email"], row["last_name"], row["normalized_title"],
                row["school"], row["department"], row["homepage"], row["direction"],
                row["title_raw"], row["title_translated"], row["title_language"],
                row["translation_status"], row["translation_engine"],
                row["classification_rules_version"], _evidence_urls(row["evidence_json"]),
                row["source_url"],
            ),
        )
        for column in (7, 16):
            cell = sheet.cell(row=sheet.max_row, column=column)
            if normalize_url(str(cell.value or "")):
                cell.hyperlink = str(cell.value)
                cell.font = Font(color="0563C1", underline="single")
    sheet.freeze_panes = "A2"
    workbook.save(path)


def _validate_accepted(rows: list[Any]) -> None:
    seen_email: set[str] = set()
    seen_name_school: set[tuple[str, str]] = set()
    seen_homepage: set[tuple[str, str]] = set()
    required = ("name", "email", "school")
    for row in rows:
        missing = [field for field in required if not str(row[field] or "").strip()]
        if missing:
            raise ValueError(f"Accepted candidate {row['id']} is missing: {', '.join(missing)}")
        email = normalize_email(row["email"])
        person_key = (str(row["name"]).casefold(), str(row["school"]).casefold())
        normalized_homepage = normalize_url(row["homepage"])
        homepage_key = (str(row["school"]).casefold(), normalized_homepage)
        if email in seen_email:
            raise ValueError(f"Duplicate accepted email: {row['email']}")
        if person_key in seen_name_school:
            raise ValueError(f"Duplicate accepted name and school: {row['name']} / {row['school']}")
        if normalized_homepage and homepage_key in seen_homepage:
            raise ValueError(f"Duplicate accepted homepage at school: {row['homepage']}")
        seen_email.add(email)
        seen_name_school.add(person_key)
        if normalized_homepage:
            seen_homepage.add(homepage_key)


def _build_audit(database: WorkflowDatabase, task_id: str, accepted: list[Any], review: list[Any]) -> dict[str, Any]:
    summary = database.summary(task_id)
    reasons = Counter()
    for row in database.list_candidates(task_id):
        reasons.update(part for part in str(row["review_reason"] or "").split(";") if part)
    sources = database.list_sources(task_id)
    generations = []
    for row in database.list_review_generations(task_id):
        item = dict(row)
        item["superseded_candidate_ids"] = json.loads(item["superseded_candidate_ids"] or "[]")
        item["requeued_school_ids"] = json.loads(item["requeued_school_ids"] or "[]")
        item["summary_json"] = json.loads(item["summary_json"] or "{}")
        generations.append(item)
    source_types = Counter(str(row["source_type"] or "unknown") for row in sources)
    fetch_states = Counter(str(row["fetch_state"] or "unknown") for row in sources)
    stop_reasons = Counter(str(row["stop_reason"]) for row in sources if row["stop_reason"])
    all_candidates = database.list_candidates(task_id)
    school_coverage = []
    for school in database.list_schools(task_id):
        school_id = int(school["id"])
        school_candidates = [row for row in all_candidates if int(row["school_id"]) == school_id]
        active = [row for row in school_candidates if row["status"] in {"accepted", "review", "candidate"}]
        school_sources = [row for row in sources if int(row["school_id"]) == school_id]
        status_counts = Counter(str(row["status"]) for row in school_candidates)
        school_stop_reasons = sorted({str(row["stop_reason"]) for row in school_sources if row["stop_reason"]})
        failed_sources = sum(
            1 for row in school_sources
            if row["fetch_state"] == "failed" or row["failure_reason"]
        )
        school_coverage.append({
            "school_id": school_id,
            "school": school["name"],
            "school_status": school["status"],
            "directory_baseline_active": len(active),
            "active_unique_people": len({str(row["normalized_person_identity"] or normalize_url(row["homepage"]) or normalize_key(row["name"])) for row in active}),
            "active_duplicate_count": 0,
            "candidate_statuses": dict(status_counts),
            "visited_sources": sum(1 for row in school_sources if row["fetch_state"] == "fetched"),
            "failed_sources": failed_sources,
            "discovery_stop_reasons": school_stop_reasons,
            "coverage_incomplete": bool(failed_sources or school_stop_reasons),
        })
    return {
        "task_id": task_id,
        "discipline": summary["discipline"],
        "status": summary["status"],
        "budget_usd": summary["budget_usd"],
        "estimated_spent_usd": summary["spent_usd"],
        "schools": summary["schools"],
        "candidates": summary["candidates"],
        "historical_people": summary["historical_people"],
        "accepted_rows": len(accepted),
        "review_rows": len(review),
        "rejection_and_review_reasons": dict(reasons.most_common()),
        "reprocessing_generations": generations,
        "source_summary": {
            "total": len(sources),
            "by_type": dict(source_types),
            "by_fetch_state": dict(fetch_states),
            "stop_reasons": dict(stop_reasons),
        },
        "completed_preservation": {
            "accepted_rows": len(accepted),
            "accepted_ids": [int(row["id"]) for row in accepted],
            "superseded_accepted_ids": [],
        },
        "school_coverage": school_coverage,
    }


def _evidence_urls(value: str) -> str:
    try:
        evidence = json.loads(value or "[]")
    except (TypeError, json.JSONDecodeError):
        return ""
    urls = dict.fromkeys(
        str(item.get("source_url") or "")
        for item in evidence
        if isinstance(item, dict) and normalize_url(str(item.get("source_url") or ""))
    )
    return "\n".join(urls)


def _append_text_row(sheet: Any, values: Iterable[Any]) -> None:
    materialized = ["" if value is None else str(value) for value in values]
    sheet.append(materialized)
    for cell in sheet[sheet.max_row]:
        cell.data_type = "s"


def _verify_final_workbook(path: Path, expected_rows: int) -> None:
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != ["学者信息"]:
        raise ValueError("Final workbook must contain exactly one worksheet named 学者信息")
    sheet = workbook.active
    if tuple(cell.value for cell in sheet[1]) != FINAL_HEADERS:
        raise ValueError("Final workbook column order is invalid")
    if sheet.max_row != expected_rows + 1:
        raise ValueError("Final workbook row count does not match accepted candidates")
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=7).value and not sheet.cell(row=row, column=7).hyperlink:
            raise ValueError(f"Homepage is not a hyperlink at row {row}")
        for column in (1, 3, 4, 5, 6, 8):
            cell = sheet.cell(row=row, column=column)
            # Excel serializes empty strings as blank cells (data_type "n").
            # Non-empty web data must still round-trip strictly as text.
            if cell.value is not None and cell.data_type != "s":
                raise ValueError(f"Expected text cell at row {row}, column {column}")
    workbook.close()
