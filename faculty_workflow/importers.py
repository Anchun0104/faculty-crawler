from __future__ import annotations

import csv
import json
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

from openpyxl import load_workbook

from faculty_workflow.database import WorkflowDatabase
from faculty_workflow.models import SchoolInput, normalize_key


SCHOOL_ALIASES = ("school", "university", "university_en", "学校", "工作单位（学校）", "工作单位")
ROW_ALIASES = ("row", "original_row", "编号", "原始编号")
DOMAIN_ALIASES = ("official_domain", "domain", "官网域名")
DIRECTORY_ALIASES = ("directory_url", "faculty_url", "教师目录", "目录网址")
NAME_ALIASES = ("name", "full_name", "全名", "姓名")
EMAIL_ALIASES = ("email", "邮箱")
HOMEPAGE_ALIASES = ("homepage", "profile_url", "profile url", "个人主页")


@dataclass(frozen=True)
class HistoryImportSummary:
    people_rows: int = 0
    processed_schools: int = 0


def load_schools(path: str | Path) -> list[SchoolInput]:
    rows = _load_rows(Path(path))
    if not rows:
        raise ValueError("School file contains no data rows")
    schools: list[SchoolInput] = []
    seen: set[str] = set()
    for row in rows:
        name = _value(row, SCHOOL_ALIASES)
        if not name:
            continue
        key = normalize_key(name)
        if key in seen:
            raise ValueError(f"Duplicate school in input: {name}")
        seen.add(key)
        directory_url = _value(row, DIRECTORY_ALIASES)
        if not directory_url:
            raise ValueError(
                f"directory_url is required for {name}; AI does not search for or guess directory URLs"
            )
        official_domain = _value(row, DOMAIN_ALIASES).lower()
        if not official_domain and directory_url:
            official_domain = (urlparse(directory_url).hostname or "").lower()
        schools.append(
            SchoolInput(
                name=name,
                original_row=_value(row, ROW_ALIASES),
                official_domain=official_domain,
                directory_url=directory_url,
            )
        )
    if not schools:
        raise ValueError(f"No school-name column found; supported headers: {SCHOOL_ALIASES}")
    return schools


def import_history(
    database: WorkflowDatabase,
    task_id: str,
    paths: Iterable[str | Path],
) -> HistoryImportSummary:
    people_rows = 0
    for raw_path in paths:
        path = Path(raw_path)
        for row in _load_rows(path):
            database.add_historical_person(
                task_id,
                email=_value(row, EMAIL_ALIASES),
                name=_value(row, NAME_ALIASES),
                school=_value(row, SCHOOL_ALIASES),
                homepage=_value(row, HOMEPAGE_ALIASES),
                source_file=path.name,
            )
            people_rows += 1
    return HistoryImportSummary(people_rows=people_rows)


def import_processed_schools(
    database: WorkflowDatabase,
    task_id: str,
    paths: Iterable[str | Path],
) -> HistoryImportSummary:
    count = 0
    for raw_path in paths:
        path = Path(raw_path)
        if path.suffix.casefold() == ".txt":
            names = [line.strip() for line in path.read_text(encoding="utf-8-sig").splitlines() if line.strip()]
        else:
            names = [_value(row, SCHOOL_ALIASES) for row in _load_rows(path)]
        for name in names:
            if name:
                database.add_processed_school(task_id, name, path.name)
                count += 1
    return HistoryImportSummary(processed_schools=count)


def _load_rows(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        raise FileNotFoundError(path)
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [_clean_row(row) for row in csv.DictReader(handle)]
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        rows: list[dict[str, str]] = []
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers = [str(value or "").strip() for value in next(iterator)]
            except StopIteration:
                continue
            for values in iterator:
                row = {headers[index]: values[index] for index in range(min(len(headers), len(values))) if headers[index]}
                if any(value not in (None, "") for value in row.values()):
                    rows.append(_clean_row(row))
        workbook.close()
        return rows
    if suffix == ".json":
        raw = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(raw, dict):
            raw = raw.get("records") or raw.get("data") or raw.get("rows") or []
        if not isinstance(raw, list):
            raise ValueError(f"JSON input must contain a list: {path}")
        return [_clean_row(dict(item)) for item in raw if isinstance(item, dict)]
    if suffix == ".txt":
        return [{"school": line.strip()} for line in path.read_text(encoding="utf-8-sig").splitlines() if line.strip()]
    raise ValueError(f"Unsupported input format: {path.suffix}")


def _clean_row(row: dict[Any, Any]) -> dict[str, str]:
    return {str(key or "").strip(): str(value or "").strip() for key, value in row.items() if key is not None}


def _value(row: dict[str, str], aliases: tuple[str, ...]) -> str:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(_normalize_header(alias), "").strip()
        if value:
            return value
    return ""


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or "")).casefold()
    return "".join(char for char in normalized if char.isalnum())
