import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from crawler.diagnostics_export import EXPORT_COLUMNS, export_records
from crawler.parsers import FacultyRecord


class DiagnosticsExportTests(unittest.TestCase):
    def setUp(self):
        self.record = FacultyRecord(
            name="Élodie 张",
            title="Assegnista di ricerca",
            profile_url="https://example.edu/people/elodie",
            email="elodie@example.edu",
            title_translated="Research manager",
            title_language="it",
            staff_classification="review",
            academic_track="research",
            affiliation_status="current",
            classification_reason="translated_title_review",
            matched_rule="",
            confidence_tier="low",
            translation_status="translated",
            translation_engine="LibreTranslate",
            classification_rules_version="2026.07.2",
            source_url="https://example.edu/faculty",
        )

    def test_csv_writes_stable_columns_and_unicode_values(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "review.csv"
            result = export_records([self.record], path)

            self.assertEqual(result, path)
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.DictReader(stream))
            self.assertEqual(list(rows[0]), list(EXPORT_COLUMNS))
            self.assertEqual(rows[0]["name"], "Élodie 张")
            self.assertEqual(rows[0]["title_translated"], "Research manager")

    def test_empty_csv_still_writes_header(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "empty.csv"
            export_records([], path)
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                self.assertEqual(next(csv.reader(stream)), list(EXPORT_COLUMNS))

    def test_xlsx_writes_header_and_values(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "review.xlsx"
            export_records([self.record], path)
            workbook = load_workbook(path, read_only=True)
            try:
                sheet = workbook.active
                rows = list(sheet.iter_rows(values_only=True))
            finally:
                workbook.close()
            self.assertEqual(list(rows[0]), list(EXPORT_COLUMNS))
            self.assertEqual(rows[1][0], "Élodie 张")
            self.assertEqual(rows[1][1], "Assegnista di ricerca")

    def test_explicit_format_overrides_extension(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "review.output"
            export_records([self.record], path, format="csv")
            self.assertTrue(path.read_bytes().startswith(b"\xef\xbb\xbf"))

    def test_unsupported_format_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaises(ValueError):
                export_records([self.record], Path(directory) / "review.json")


if __name__ == "__main__":
    unittest.main()
