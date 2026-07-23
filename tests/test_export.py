import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from crawler.faculty_crawler import (
    FacultyRecord,
    TitlePendingRecord,
    export_title_pending_to_excel,
    export_to_excel,
)


class ExportTests(unittest.TestCase):
    def test_export_writes_expected_excel_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = f"{temp_dir}/faculty.xlsx"
            records = [
                FacultyRecord(
                    name="Radia Perlman",
                    title="Professor",
                    profile_url="https://example.edu/people/radia-perlman/",
                )
            ]

            export_to_excel(records, output_path)

            workbook = load_workbook(output_path)
            sheet = workbook.active
            self.assertEqual([cell.value for cell in sheet[1]], ["Name", "Title", "Profile_URL", "Email"])
            self.assertEqual(
                [cell.value for cell in sheet[2]],
                [
                    "Radia Perlman",
                    "Professor",
                    "https://example.edu/people/radia-perlman/",
                    None,
                ],
            )

    def test_pending_export_uses_derived_path_columns_overwrite_and_normalized_url_deduplication(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_output = Path(temp_dir) / "output" / "murdoch_psycho_faculty.xlsx"
            duplicate_records = [
                TitlePendingRecord(
                    name="Ada Lovelace",
                    directory_title="Dr.",
                    profile_url="https://example.edu/people/ada/?utm_source=directory",
                    email="ada@example.edu",
                    section="Academic Staff",
                    source_url="https://example.edu/faculty",
                    pending_reason="honorific_only_title",
                ),
                TitlePendingRecord(
                    name="Ada Lovelace",
                    directory_title="",
                    profile_url="https://example.edu/people/ada/",
                    section="Academic Staff",
                    source_url="https://example.edu/faculty?page=2",
                    pending_reason="missing_title",
                ),
            ]

            pending_path = export_title_pending_to_excel(duplicate_records, main_output)

            self.assertEqual(
                pending_path,
                main_output.parent / "pending_title" / "murdoch_psycho_faculty_title_pending.xlsx",
            )
            workbook = load_workbook(pending_path)
            sheet = workbook.active
            self.assertEqual(
                [cell.value for cell in sheet[1]],
                [
                    "Name",
                    "Directory_Title",
                    "Profile_URL",
                    "Email",
                    "Section",
                    "Source_URL",
                    "Pending_Reason",
                    "Next_Action",
                    "Status",
                ],
            )
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual([cell.value for cell in sheet[2]], [
                "Ada Lovelace",
                "Dr.",
                "https://example.edu/people/ada/?utm_source=directory",
                "ada@example.edu",
                "Academic Staff",
                "https://example.edu/faculty",
                "honorific_only_title",
                "extract_title_from_profile",
                "pending",
            ])

            replacement = TitlePendingRecord(
                name="Grace Hopper",
                directory_title="",
                profile_url="https://example.edu/people/grace",
                section="Faculty",
                source_url="https://example.edu/faculty",
                pending_reason="missing_title",
            )
            export_title_pending_to_excel([replacement], main_output)
            workbook = load_workbook(pending_path)
            self.assertEqual(workbook.active.max_row, 2)
            self.assertEqual(workbook.active["A2"].value, "Grace Hopper")

    def test_pending_export_does_not_create_empty_file_and_removes_stale_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_output = Path(temp_dir) / "faculty.xlsx"
            pending_path = main_output.parent / "pending_title" / "faculty_title_pending.xlsx"
            pending_path.parent.mkdir(parents=True)
            pending_path.write_bytes(b"stale")

            result = export_title_pending_to_excel([], main_output)

            self.assertIsNone(result)
            self.assertFalse(pending_path.exists())


if __name__ == "__main__":
    unittest.main()
