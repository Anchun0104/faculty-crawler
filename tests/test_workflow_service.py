from __future__ import annotations

import json
import gzip
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from crawler.parsers import FacultyRecord
from crawler.title_classifier import TitleClassifier
from crawler.title_pipeline import TitlePipeline
from faculty_workflow.database import WorkflowDatabase
from faculty_workflow.discovery import DiscoveryHint, EmptyDiscoveryProvider
from faculty_workflow.fetcher import AccessBlockedError, FetchError, FetchPolicy, FetchedPage, PageFetcher, html_to_text, retry_delay_seconds, url_is_on_domain
from faculty_workflow.models import CandidateExtraction, DisciplinePolicy, SchoolInput
from faculty_workflow.providers import DeepSeekProvider, MissingAPIKeyError, ProviderResult
from faculty_workflow.service import (
    BudgetExceededError,
    WorkflowService,
    _apply_title_pipeline,
    _apply_relevance_rule,
    _ground_extraction,
    _hard_exclusion,
    _local_profile_extraction,
    _local_last_name,
    _load_cached_source_page,
    _merge_directory_card_evidence,
    _needs_escalation,
    _profile_excerpt,
    _store_directory_seed,
)
from tests.test_workflow_pdf_documents import make_text_pdf


class FakeProvider:
    def estimate_request_cost(self, model, input_characters, max_output_tokens):
        return 0.01

    def generate_policy(self, discipline, model):
        raise AssertionError("not used")

    def discover_sources(self, school, policy, model, max_results=10):
        raise AssertionError("seed URL should avoid discovery")

    def extract_profile(self, **kwargs):
        source = kwargs["profile_url"]
        quotes = {
            "name": "Ada Lovelace",
            "email": "ada@example.edu",
            "title": "Professor of Physics",
            "department": "Department of Physics",
            "professional_relevance": "Physics",
        }
        data = {
            "name": "Ada Lovelace",
            "email": "ada@example.edu",
            "last_name": "Lovelace",
            "title_raw": "Professor of Physics",
            "normalized_title": "Professor",
            "department": "Department of Physics",
            "homepage": source,
            "professional_relevance": "relevant",
            "email_ownership": "verified",
            "homepage_identity": "verified",
            "official_source": True,
            "group_homepage": False,
            "evidence": [
                {"field": field, "quote": quote, "source_url": source, "extraction_method": "model", "status": "supported"}
                for field, quote in quotes.items()
            ],
            "failure_reasons": [],
        }
        return ProviderResult(data, kwargs["model"], "resp", 100, 100, 0, 0.0007)


class FakeFetcher:
    def __init__(self, root: Path):
        self.root = root

    def fetch(self, url, snapshot_dir, *, expand_directory=False):
        path = self.root / ("profile.html.gz" if url.endswith("/ada") else "directory.html.gz")
        path.write_bytes(b"snapshot")
        html = "<html><title>Ada</title><body>Ada Lovelace Professor of Physics, Department of Physics, ada@example.edu</body></html>"
        return FetchedPage(url, url, 200, "Ada", html, html_to_text(html), "hash", path)


class FakeCrawler:
    title_pending_records = []

    def __init__(self, timeout):
        pass

    def crawl(self, url):
        raise AssertionError("workflow should parse the already fetched directory snapshot")

    def parse_fetched_directory(self, url, html, fetch_status):
        self.parsed_url = url
        self.parsed_html = html
        self.parsed_status = fetch_status
        return [FacultyRecord("Ada Lovelace", "Professor of Physics", "https://example.edu/ada", "ada@example.edu")]


class WorkflowServiceTests(unittest.TestCase):
    def test_direct_urls_create_confirmed_generic_task_without_ai(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            database = WorkflowDatabase(Path(temp_dir) / "workflow.db")
            service = WorkflowService(database, provider=FakeProvider())

            task_id = service.create_direct_url_task(
                directory_urls=["https://www.example.edu/people", "https://law.example.edu/faculty"],
                output_dir=Path(temp_dir) / "output",
            )

            task = database.get_task(task_id)
            schools = database.list_schools(task_id)
            self.assertEqual(task["discipline"], "General Faculty")
            self.assertEqual(task["routine_model"], "local-only")
            self.assertTrue(task["policy_confirmed"])
            self.assertEqual(
                [(row["name"], row["directory_url"]) for row in schools],
                [
                    ("example.edu", "https://www.example.edu/people"),
                    ("law.example.edu", "https://law.example.edu/faculty"),
                ],
            )

    def test_user_verified_external_directory_is_collected_with_primary_domain_email(self) -> None:
        class ExternalDirectoryCrawler(FakeCrawler):
            def parse_fetched_directory(self, url, html, fetch_status):
                return [FacultyRecord(
                    "Ada Lovelace",
                    "Professor of Physics",
                    "https://research-institute.org/ada",
                    "ada@university.edu",
                )]

        class ExternalDirectoryFetcher(FakeFetcher):
            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                result = super().fetch(url, snapshot_dir, expand_directory=expand_directory)
                if url.endswith("/faculty"):
                    html = (
                        "<html><title>Research Institute Faculty</title><body>"
                        "Ada Lovelace Professor of Physics ada@university.edu"
                        "</body></html>"
                    )
                    return FetchedPage(url, url, 200, "Research Institute Faculty", html, html_to_text(html), "directory", result.snapshot_path)
                return result

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput(
                    "Example University",
                    official_domain="university.edu",
                    directory_url="https://research-institute.org/faculty",
                )],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )

            summary = WorkflowService(
                database,
                provider=FakeProvider(),
                fetcher=ExternalDirectoryFetcher(root),
                crawler_factory=lambda timeout: ExternalDirectoryCrawler(timeout),
            ).run_task(task_id)

            self.assertEqual(summary["candidates"], {"accepted": 1})
            self.assertEqual(database.list_schools(task_id)[0]["status"], "completed")

    def test_persisted_pdf_snapshot_is_rehydrated(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            snapshot = Path(temp_dir) / "profile.pdf"
            snapshot.write_bytes(
                make_text_pdf("Ada Lovelace Professor ada@example.edu", title="Ada PDF")
            )
            source = {
                "fetch_state": "fetched",
                "http_status": 200,
                "content_hash": "pdf-hash",
                "snapshot_path": str(snapshot),
                "url": "https://example.edu/ada.pdf",
                "final_url": "https://example.edu/ada.pdf",
            }

            page = _load_cached_source_page(source, "https://example.edu/ada.pdf")

            self.assertIsNotNone(page)
            self.assertEqual(page.title, "Ada PDF")
            self.assertIn("ada@example.edu", page.text)
            self.assertEqual(page.html, "")

    def test_review_generation_does_not_run_unrelated_failed_school(self) -> None:
        class RecordingService(WorkflowService):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self.processed_school_names = []

            def _run_school(self, task_id, school, policy, on_progress) -> None:
                self.processed_school_names.append(str(school["name"]))
                self.database.update_school(school["id"], status="completed")

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [
                    SchoolInput("Reviewed University", official_domain="reviewed.example.edu"),
                    SchoolInput("Unrelated Failed University", official_domain="failed.example.edu"),
                ],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )
            reviewed_school, failed_school = database.list_schools(task_id)
            database.add_candidate(
                task_id,
                reviewed_school["id"],
                CandidateExtraction(name="Ada Lovelace", homepage="https://reviewed.example.edu/ada"),
                direction="Physics",
                source_url="https://reviewed.example.edu/faculty",
                status="review",
            )
            database.update_school(reviewed_school["id"], status="review")
            database.update_school(failed_school["id"], status="failed", failure_reason="transient failure")

            service = RecordingService(database, provider=FakeProvider())
            service.run_review_generation(task_id)

            self.assertEqual(service.processed_school_names, ["Reviewed University"])

    def test_complete_official_directory_evidence_skips_profile_fetch(self) -> None:
        class DirectoryCompleteFetcher(FakeFetcher):
            def __init__(self, root):
                super().__init__(root)
                self.profile_requests = 0

            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                if url.endswith("/ada"):
                    self.profile_requests += 1
                    raise FetchError("profile must not be required")
                return super().fetch(url, snapshot_dir, expand_directory=expand_directory)

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput(
                    "Example University", "1", "example.edu", "https://example.edu/faculty"
                )],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )
            fetcher = DirectoryCompleteFetcher(root)

            summary = WorkflowService(
                database,
                provider=DeepSeekProvider(api_key=""),
                fetcher=fetcher,
                crawler_factory=lambda timeout: FakeCrawler(timeout),
            ).run_task(task_id)

            self.assertEqual(summary["candidates"], {"accepted": 1})
            self.assertEqual(fetcher.profile_requests, 0)
            self.assertEqual(database.list_candidates(task_id, ["review"]), [])

    def test_cached_official_email_source_is_reused(self) -> None:
        class CachedEmailFetcher(FakeFetcher):
            def __init__(self, root):
                super().__init__(root)
                self.email_source_requests = 0

            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                if url.endswith("/contact"):
                    self.email_source_requests += 1
                    raise FetchError("cached email source must not be refetched")
                result = super().fetch(url, snapshot_dir, expand_directory=expand_directory)
                if url.endswith("/ada"):
                    html = (
                        "<html><body>Ada Lovelace Professor of Physics Department of Physics "
                        "<a href='/people/ada/contact'>Contact</a></body></html>"
                    )
                    return FetchedPage(
                        url, url, 200, "Ada Lovelace", html, html_to_text(html), "profile-hash",
                        result.snapshot_path,
                    )
                return result

        class MissingDirectoryEmailCrawler(FakeCrawler):
            def parse_fetched_directory(self, url, html, fetch_status):
                return [FacultyRecord(
                    "Ada Lovelace", "Professor of Physics", "https://example.edu/ada", ""
                )]

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput(
                    "Example University", "1", "example.edu", "https://example.edu/faculty"
                )],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )
            school_id = int(database.list_schools(task_id)[0]["id"])
            contact_html = "<html><body>Ada Lovelace ada@example.edu</body></html>"
            contact_snapshot = root / "contact.html.gz"
            with gzip.open(contact_snapshot, "wb") as handle:
                handle.write(contact_html.encode("utf-8"))
            contact_source_id = database.add_source(
                task_id,
                school_id,
                "https://example.edu/people/ada/contact",
                "official_email_source",
                official=True,
            )
            database.update_source(
                contact_source_id,
                final_url="https://example.edu/people/ada/contact",
                http_status=200,
                content_hash="contact-hash",
                snapshot_path=str(contact_snapshot),
                fetch_state="fetched",
            )
            fetcher = CachedEmailFetcher(root)

            summary = WorkflowService(
                database,
                provider=DeepSeekProvider(api_key=""),
                fetcher=fetcher,
                crawler_factory=lambda timeout: MissingDirectoryEmailCrawler(timeout),
            ).run_task(task_id)

            self.assertEqual(summary["candidates"], {"accepted": 1})
            self.assertEqual(fetcher.email_source_requests, 0)
            self.assertEqual(database.list_candidates(task_id, ["accepted"])[0]["email"], "ada@example.edu")

    def test_service_accepts_local_title_and_url_only_discovery_dependencies(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            database = WorkflowDatabase(Path(temp_dir) / "workflow.db")
            title_pipeline = TitlePipeline(TitleClassifier())
            discovery_provider = EmptyDiscoveryProvider()

            service = WorkflowService(
                database,
                provider=FakeProvider(),
                title_pipeline=title_pipeline,
                discovery_provider=discovery_provider,
            )

            self.assertIs(service.title_pipeline, title_pipeline)
            self.assertIs(service.discovery_provider, discovery_provider)

    def test_workflow_title_pipeline_preserves_official_original(self) -> None:
        processed = TitlePipeline(TitleClassifier()).process("Professor")

        self.assertEqual(processed.title_original, "Professor")
        self.assertEqual(processed.title_translated, "")

    def test_translated_title_assists_classification_without_becoming_evidence(self) -> None:
        class FakeTitlePipeline:
            def process(self, title_original, *, language_hint=""):
                self.call = (title_original, language_hint)
                return SimpleNamespace(
                    title_original=title_original,
                    title_translated="Associate Professor",
                    title_language="de",
                    translation_status="translation_success",
                    translation_engine="argos",
                    rules_version="2026.07.2",
                    classification=SimpleNamespace(classification=SimpleNamespace(value="include")),
                )

        pipeline = FakeTitlePipeline()
        extraction = CandidateExtraction(
            name="Ada Lovelace",
            title_raw="Universitätsprofessorin",
            evidence=(
                {
                    "field": "title",
                    "quote": "Universitätsprofessorin",
                    "source_url": "https://example.edu/ada",
                    "extraction_method": "directory_card",
                    "status": "supported",
                },
            ),
        )
        extraction = CandidateExtraction.from_mapping({
            **extraction.__dict__,
            "evidence": list(extraction.evidence),
        })

        enriched = _apply_title_pipeline(
            extraction,
            pipeline,
            DisciplinePolicy("Physics", ("physics",), ()),
            language_hint="de-DE",
        )

        self.assertEqual(pipeline.call, ("Universitätsprofessorin", "de-DE"))
        self.assertEqual(enriched.title_raw, "Universitätsprofessorin")
        self.assertEqual(enriched.title_translated, "Associate Professor")
        self.assertEqual(enriched.normalized_title, "Associate Professor")
        self.assertEqual(enriched.title_language, "de")
        self.assertEqual(enriched.translation_engine, "argos")
        self.assertEqual(enriched.classification_rules_version, "2026.07.2")
        self.assertEqual([item.quote for item in enriched.evidence], ["Universitätsprofessorin"])

    def test_offline_end_to_end_run_and_export(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            policy = DisciplinePolicy("Physics", ("physics",), ())
            task_id = database.create_task(
                policy,
                [SchoolInput("Example University", "1", "example.edu", "https://example.edu/faculty")],
                output_dir=root / "output",
                policy_confirmed=True,
            )
            service = WorkflowService(
                database,
                provider=FakeProvider(),
                fetcher=FakeFetcher(root),
                crawler_factory=lambda timeout: FakeCrawler(timeout),
            )

            summary = service.run_task(task_id)
            paths = service.export(task_id)

            self.assertEqual(summary["status"], "completed")
            self.assertEqual(summary["candidates"], {"accepted": 1})
            workbook = load_workbook(paths["final"])
            self.assertEqual(workbook.sheetnames, ["学者信息"])
            self.assertEqual(workbook.active.max_row, 2)
            self.assertEqual(workbook.active["A2"].value, "Ada Lovelace")
            self.assertIsNotNone(workbook.active["G2"].hyperlink)
            completed_workbook = load_workbook(paths["completed"])
            completed_headers = [cell.value for cell in completed_workbook.active[1]]
            self.assertEqual(tuple(completed_headers[:8]), list(workbook.active.values)[0])
            self.assertIn("Original title", completed_headers)
            self.assertIn("Translation status", completed_headers)
            self.assertIn("Evidence URLs", completed_headers)
            audit = json.loads(paths["audit"].read_text(encoding="utf-8"))
            self.assertEqual(audit["accepted_rows"], 1)
            self.assertIn("reprocessing_generations", audit)
            self.assertIn("source_summary", audit)
            self.assertEqual(audit["completed_preservation"]["accepted_rows"], 1)

    def test_local_only_task_never_requires_or_calls_a_model(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            policy = DisciplinePolicy("Physics", ("physics",), ())
            task_id = database.create_task(
                policy,
                [SchoolInput("Example University", "1", "example.edu", "https://example.edu/faculty")],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )
            service = WorkflowService(
                database,
                provider=DeepSeekProvider(api_key=""),
                fetcher=FakeFetcher(root),
                crawler_factory=lambda timeout: FakeCrawler(timeout),
            )
            summary = service.run_task(task_id)
            self.assertEqual(summary["status"], "completed")
            self.assertEqual(summary["candidates"], {"accepted": 1})
            self.assertEqual(summary["spent_usd"], 0.0)

    def test_local_only_service_selects_personal_email_over_global_contact(self) -> None:
        class NoEmailCrawler(FakeCrawler):
            def parse_fetched_directory(self, url, html, fetch_status):
                return [FacultyRecord(
                    "Ada Lovelace", "Professor of Physics", "https://example.edu/ada", ""
                )]

        class MultipleEmailFetcher(FakeFetcher):
            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                result = super().fetch(url, snapshot_dir, expand_directory=expand_directory)
                if not url.endswith("/ada"):
                    return result
                html = """<h1>Ada Lovelace</h1><p>Professor of Physics</p>
                    <p>Department of Physics</p><p>info@example.edu</p>
                    <a href='mailto:ada.lovelace@example.edu'>Email Ada</a>"""
                return FetchedPage(
                    url, url, 200, "Ada Lovelace", html, html_to_text(html), "profile-hash", result.snapshot_path
                )

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput(
                    "Example University", "1", "example.edu", "https://example.edu/faculty"
                )],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )
            summary = WorkflowService(
                database,
                provider=DeepSeekProvider(api_key=""),
                fetcher=MultipleEmailFetcher(root),
                crawler_factory=lambda timeout: NoEmailCrawler(timeout),
            ).run_task(task_id)
            self.assertEqual(summary["candidates"], {"accepted": 1})
            candidate = database.list_candidates(task_id, ["accepted"])[0]
            self.assertEqual(candidate["email"], "ada.lovelace@example.edu")
            evidence = json.loads(candidate["evidence_json"])
            self.assertTrue(any(
                item["field"] == "email" and item["status"] == "supported" for item in evidence
            ))
            self.assertIn("person_profile", {row["source_type"] for row in database.list_sources(task_id)})

    def test_directory_name_and_official_email_can_pass_without_personal_homepage(self) -> None:
        class CardFetcher(FakeFetcher):
            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                snapshot = self.root / "cards.html.gz"
                snapshot.write_bytes(b"snapshot")
                html = """
                <main><h1>Physics staff</h1>
                  <article class="person"><h3>Ada Lovelace</h3><p>Professor</p>
                    <a href="mailto:ada@example.edu">ada@example.edu</a></article>
                  <article class="person"><h3>Grace Hopper</h3><p>Senior Lecturer</p>
                    <a href="mailto:grace@example.edu">grace@example.edu</a></article>
                </main>
                """
                return FetchedPage(url, url, 200, "Physics staff", html, html_to_text(html), "cards", snapshot)

        class EmptyLegacyCrawler(FakeCrawler):
            def parse_fetched_directory(self, url, html, fetch_status):
                return []

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput(
                    "Example University", "1", "example.edu", "https://example.edu/physics/staff"
                )],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )
            summary = WorkflowService(
                database,
                provider=DeepSeekProvider(api_key=""),
                fetcher=CardFetcher(root),
                crawler_factory=lambda timeout: EmptyLegacyCrawler(timeout),
            ).run_task(task_id)

            self.assertEqual(summary["candidates"], {"accepted": 2})
            rows = database.list_candidates(task_id, ["accepted"])
            self.assertEqual({row["email"] for row in rows}, {"ada@example.edu", "grace@example.edu"})
            self.assertTrue(all(not row["homepage"] for row in rows))

    def test_exact_name_merges_secondary_official_email_evidence(self) -> None:
        seeds = {}
        _store_directory_seed(seeds, {
            "name": "Ada Lovelace",
            "title": "Professor",
            "profile_url": "https://example.edu/people/ada",
            "email": "",
            "directory_url": "https://example.edu/physics/faculty",
            "directory_quote": "Ada Lovelace | Professor",
            "extraction_method": "directory_card",
        })
        _store_directory_seed(seeds, {
            "name": "Ada Lovelace",
            "title": "",
            "profile_url": "https://example.edu/quantum-lab/ada",
            "email": "ada.lovelace@example.edu",
            "directory_url": "https://example.edu/quantum-lab/team",
            "directory_quote": "Ada Lovelace | ada.lovelace@example.edu",
            "extraction_method": "research_unit_card",
        })

        self.assertEqual(len(seeds), 1)
        seed = next(iter(seeds.values()))
        self.assertEqual(seed["profile_url"], "https://example.edu/people/ada")
        self.assertEqual(seed["email"], "ada.lovelace@example.edu")
        extraction = _merge_directory_card_evidence(
            CandidateExtraction(
                name="Ada Lovelace", homepage="https://example.edu/people/ada",
                official_source=True,
            ),
            seed,
            DisciplinePolicy("Physics", ("physics",), ()),
        )
        email_evidence = [item for item in extraction.evidence if item.field == "email"]
        self.assertEqual(extraction.email, "ada.lovelace@example.edu")
        self.assertEqual(extraction.email_ownership, "verified")
        self.assertEqual(len(email_evidence), 1)
        self.assertEqual(email_evidence[0].source_url, "https://example.edu/quantum-lab/team")
        self.assertIn(extraction.email, email_evidence[0].quote)

    def test_different_people_sharing_one_lab_pdf_remain_distinct_seeds(self) -> None:
        seeds = {}
        shared_pdf = "https://example.edu/labs/quantum.pdf"
        _store_directory_seed(seeds, {
            "name": "Ada Lovelace",
            "title": "Professor",
            "profile_url": shared_pdf,
            "email": "",
            "directory_url": "https://example.edu/physics/faculty",
            "directory_quote": "Ada Lovelace | Professor",
            "extraction_method": "directory_card",
        })
        _store_directory_seed(seeds, {
            "name": "Grace Hopper",
            "title": "Associate Professor",
            "profile_url": shared_pdf,
            "email": "",
            "directory_url": "https://example.edu/physics/faculty",
            "directory_quote": "Grace Hopper | Associate Professor",
            "extraction_method": "directory_card",
        })

        self.assertEqual(len(seeds), 2)
        self.assertEqual(
            {seed["name"] for seed in seeds.values()},
            {"Ada Lovelace", "Grace Hopper"},
        )

    def test_laboratory_label_is_not_stored_as_a_person_seed(self) -> None:
        seeds = {}

        _store_directory_seed(seeds, {
            "name": "QG Lab",
            "title": "Designated Assistant Professor",
            "profile_url": "https://example.edu/labs/quantum.pdf",
            "email": "",
            "directory_url": "https://example.edu/physics/faculty",
            "directory_quote": "QG Lab | Designated Assistant Professor",
            "extraction_method": "directory_card",
        })

        self.assertEqual(seeds, {})

    def test_profile_fetch_failure_is_reviewed_without_failing_the_school(self) -> None:
        class FailingProfileFetcher(FakeFetcher):
            def __init__(self, root):
                super().__init__(root)
                self.profile_policies = []

            def fetch(self, url, snapshot_dir, *, expand_directory=False, policy=None):
                if url.endswith("/ada"):
                    self.profile_policies.append(policy)
                    raise FetchError("profile timed out")
                return super().fetch(url, snapshot_dir, expand_directory=expand_directory)

        class IncompleteDirectoryCrawler(FakeCrawler):
            def parse_fetched_directory(self, url, html, fetch_status):
                return [FacultyRecord(
                    "Ada Lovelace", "Professor of Physics", "https://example.edu/ada", ""
                )]

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput("Example University", "1", "example.edu", "https://example.edu/faculty")],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )
            fetcher = FailingProfileFetcher(root)
            summary = WorkflowService(
                database,
                provider=DeepSeekProvider(api_key=""),
                fetcher=fetcher,
                crawler_factory=lambda timeout: IncompleteDirectoryCrawler(timeout),
            ).run_task(task_id)
            self.assertEqual(summary["status"], "completed")
            self.assertEqual(summary["schools"], {"review": 1})
            candidate = database.list_candidates(task_id, ["review"])[0]
            self.assertIn("profile_fetch_failed", candidate["review_reason"])
            self.assertEqual(fetcher.profile_policies, [FetchPolicy.person_profile()])

    def test_budget_reservation_pauses_before_call_and_can_resume(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            database = WorkflowDatabase(Path(temp_dir) / "workflow.db")
            policy = DisciplinePolicy("Physics", ("physics",), ())
            task_id = database.create_task(policy, [SchoolInput("Example University")], output_dir=temp_dir, budget_usd=0.005, policy_confirmed=True)
            service = WorkflowService(database, provider=FakeProvider())
            called = []
            with self.assertRaises(BudgetExceededError):
                service._provider_call(
                    task_id,
                    operation="test",
                    model="deepseek-v4-flash",
                    input_characters=10,
                    max_output_tokens=10,
                    call=lambda: called.append(True),
                )
            self.assertEqual(called, [])
            self.assertEqual(database.get_task(task_id)["status"], "paused_budget")
            database.set_budget(task_id, 1.0)
            self.assertEqual(database.get_task(task_id)["status"], "ready")

    def test_domain_and_text_helpers_are_bounded(self) -> None:
        self.assertTrue(url_is_on_domain("https://people.example.edu/ada", "example.edu"))
        self.assertFalse(url_is_on_domain("https://example.edu.bad.test/ada", "example.edu"))
        self.assertEqual(html_to_text("<p>Ada</p><script>secret()</script><p>Physics</p>"), "Ada Physics")
        self.assertEqual([retry_delay_seconds(index) for index in range(5)], [1.0, 2.0, 4.0, 8.0, 8.0])

    def test_missing_api_key_fails_before_school_state_changes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            database = WorkflowDatabase(Path(temp_dir) / "workflow.db")
            policy = DisciplinePolicy("Physics", ("physics",), ())
            task_id = database.create_task(
                policy, [SchoolInput("Example University")], output_dir=temp_dir, policy_confirmed=True
            )
            service = WorkflowService(database, provider=DeepSeekProvider(api_key=""))
            with self.assertRaises(MissingAPIKeyError):
                service.run_task(task_id)
            self.assertEqual(database.list_schools(task_id)[0]["status"], "pending")

    def test_missing_directory_url_routes_school_to_review_without_model_search(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            database = WorkflowDatabase(Path(temp_dir) / "workflow.db")
            policy = DisciplinePolicy("Physics", ("physics",), ())
            task_id = database.create_task(
                policy, [SchoolInput("Example University", official_domain="example.edu")],
                output_dir=temp_dir, policy_confirmed=True,
            )
            summary = WorkflowService(database, provider=FakeProvider()).run_task(task_id)
            self.assertEqual(summary["schools"], {"review": 1})
            self.assertEqual(database.list_schools(task_id)[0]["failure_reason"], "missing_directory_url")

    def test_model_email_and_quotes_not_in_page_are_downgraded(self) -> None:
        root = Path(".")
        page = FetchedPage(
            "https://example.edu/ada", "https://example.edu/ada", 200, "Ada",
            "<p>Ada Lovelace</p>", "Ada Lovelace works in physics.", "hash", root / "snapshot.gz",
        )
        extraction = CandidateExtraction(
            name="Ada Lovelace", email="guessed@example.edu", email_ownership="verified",
            homepage_identity="verified",
            evidence=tuple(),
        )
        grounded = _ground_extraction(extraction, page)
        self.assertEqual(grounded.email_ownership, "uncertain")
        self.assertIn("email_not_present_in_profile_page", grounded.failure_reasons)

    def test_cost_controls_exclude_unambiguous_titles_and_bound_profile_text(self) -> None:
        self.assertEqual(
            _hard_exclusion({"name": "Ada", "title": "Doctoral Researcher"}),
            "directory_title_outside_confirmed_policy",
        )
        self.assertEqual(_hard_exclusion({"name": "Ada", "title": "Postdoctoral Researcher"}), "")
        self.assertEqual(_hard_exclusion({"name": "Ada", "title": "Assistant Professor"}), "")
        self.assertEqual(
            _hard_exclusion({"name": "Project Partners", "title": ""}),
            "directory_entry_is_not_person",
        )
        self.assertEqual(
            _hard_exclusion({"name": "Platform & Collaboration", "title": "Researcher"}),
            "directory_entry_is_not_person",
        )
        text = "x" * 2_000 + "Ada Lovelace Physics Department ada@example.edu " + "y" * 12_000
        excerpt = _profile_excerpt(text, {"name": "Ada Lovelace"}, max_characters=4_000)
        self.assertLessEqual(len(excerpt), 4_000)
        self.assertIn("Ada Lovelace", excerpt)

    def test_directory_card_evidence_supplies_rank_and_email_for_linked_profile(self) -> None:
        policy = DisciplinePolicy(
            "Physics", ("physics",), (),
            allowed_titles=("Postdoctoral Researcher", "Senior Researcher"),
            title_mappings={"Postdoctoral Researcher": "Postdoctoral Researcher"},
        )
        seed = {
            "name": "Ada Lovelace",
            "title": "Postdoctoral Researcher",
            "email": "ada@example.edu",
            "profile_url": "https://example.edu/ada",
            "directory_url": "https://example.edu/physics/people",
            "directory_quote": "Ada Lovelace | Postdoctoral Researcher | ada@example.edu",
        }
        merged = _merge_directory_card_evidence(
            CandidateExtraction(homepage="https://example.edu/ada"), seed, policy
        )
        self.assertEqual(merged.normalized_title, "Postdoctoral Researcher")
        self.assertEqual(merged.email, "ada@example.edu")
        self.assertEqual(merged.email_ownership, "verified")
        self.assertEqual(merged.homepage_identity, "verified")
        self.assertEqual({item.field for item in merged.evidence}, {"name", "title", "email"})

    def test_explicit_discipline_term_overrides_model_relevance_uncertainty(self) -> None:
        page = FetchedPage(
            "https://example.edu/ada", "https://example.edu/ada", 200, "Physics people",
            "<p>Department of Physics</p>", "Department of Physics", "hash", Path("snapshot.gz"),
        )
        extraction = _apply_relevance_rule(
            CandidateExtraction(professional_relevance="not_relevant"),
            page,
            DisciplinePolicy("Physics", ("quantum physics",), ()),
        )
        self.assertEqual(extraction.professional_relevance, "relevant")
        self.assertTrue(any(item.field == "professional_relevance" for item in extraction.evidence))

    def test_local_profile_extraction_recovers_mapped_title_from_profile_text(self) -> None:
        page = FetchedPage(
            "https://example.edu/ada", "https://example.edu/ada", 200, "Physics",
            "", "Ada Lovelace, Postdoc, Department of Physics", "hash", Path("snapshot.gz"),
        )
        policy = DisciplinePolicy(
            "Physics", ("physics",), (),
            allowed_titles=("Postdoctoral Researcher",),
            title_mappings={"Postdoc": "Postdoctoral Researcher"},
        )
        extraction = _local_profile_extraction(page, {"name": "Ada Lovelace"}, policy)
        self.assertEqual(extraction.title_raw, "Postdoc")
        self.assertEqual(extraction.normalized_title, "Postdoctoral Researcher")
        self.assertEqual(extraction.department, "Department of Physics")
        self.assertTrue(any(item.field == "title" for item in extraction.evidence))

    def test_local_profile_extraction_removes_appended_role_from_name(self) -> None:
        page = FetchedPage(
            "https://example.edu/atif", "https://example.edu/atif", 200, "Atif Ghafoor",
            "", "Atif Ghafoor Postdoctoral Researcher Physics", "hash", Path("snapshot.gz"),
        )
        extraction = _local_profile_extraction(
            page,
            {"name": "Atif Ghafoor Postdoctoral Researcher"},
            DisciplinePolicy("Physics", ("physics",), ()),
        )
        self.assertEqual(extraction.name, "Atif Ghafoor")
        self.assertEqual(extraction.last_name, "Ghafoor")
        self.assertEqual(_local_last_name("Zscherp, Mario"), "Zscherp")

    def test_local_profile_extraction_prefers_identity_printed_on_person_page(self) -> None:
        html = """
        <main>
          <h1 itemprop="name">Susanne Aalto</h1>
          <div class="profile-title">Full Professor, Physics and Astronomy</div>
          <a href="mailto:susanne.aalto@example.edu">susanne.aalto@example.edu</a>
        </main>
        """
        page = FetchedPage(
            "https://example.edu/persons/saalto/",
            "https://example.edu/persons/saalto/",
            200,
            "Susanne Aalto",
            html,
            "Susanne Aalto Full Professor, Physics and Astronomy susanne.aalto@example.edu",
            "hash",
            Path("snapshot.gz"),
        )

        extraction = _local_profile_extraction(
            page,
            {"name": "Highlighted programs", "title": "Full Professor"},
            DisciplinePolicy(
                "Physics",
                ("physics",),
                (),
                allowed_titles=("Professor",),
                title_mappings={"Full Professor": "Professor"},
            ),
        )

        self.assertEqual(extraction.name, "Susanne Aalto")
        self.assertEqual(extraction.email, "susanne.aalto@example.edu")
        self.assertEqual(extraction.homepage_identity, "verified")

    def test_local_physics_alias_and_german_institute_are_grounded(self) -> None:
        page = FetchedPage(
            "https://example.edu/ada", "https://example.edu/ada", 200, "I. Physikalisches Institut",
            "", "Ada Lovelace, Postdoc, I. Physikalisches Institut", "hash", Path("snapshot.gz"),
        )
        policy = DisciplinePolicy(
            "Physics", ("physics",), (),
            allowed_titles=("Postdoctoral Researcher",),
            title_mappings={"Postdoc": "Postdoctoral Researcher"},
        )
        extraction = _local_profile_extraction(page, {"name": "Ada Lovelace"}, policy)
        extraction = _apply_relevance_rule(extraction, page, policy)
        self.assertEqual(extraction.department, "I. Physikalisches Institut")
        self.assertEqual(extraction.professional_relevance, "relevant")
        self.assertTrue(any(item.field == "department" for item in extraction.evidence))

    def test_escalation_requires_a_near_pass_professional_boundary_case(self) -> None:
        near_pass = CandidateExtraction(
            name="Ada", email="ada@example.edu", normalized_title="Professor",
            department="Physics", homepage="https://example.edu/ada",
            professional_relevance="uncertain", email_ownership="verified", homepage_identity="verified",
        )
        self.assertTrue(_needs_escalation(near_pass))
        self.assertFalse(_needs_escalation(CandidateExtraction(professional_relevance="uncertain")))
        self.assertFalse(_needs_escalation(CandidateExtraction(**{**near_pass.__dict__, "email_ownership": "uncertain"})))

    def test_dynamic_expansion_stops_after_bounded_unchanged_rounds(self) -> None:
        class FakePage:
            def __init__(self):
                self.load_actions = iter([
                    {"acted": True, "profile_count": 2},
                    {"acted": True, "profile_count": 3},
                    {"acted": True, "profile_count": 3},
                ])
                self.scroll_actions = iter([
                    {"acted": True, "profile_count": 3, "kind": "window_scroll"},
                    {"acted": True, "profile_count": 3, "kind": "window_scroll"},
                    {"acted": True, "profile_count": 3, "kind": "window_scroll"},
                ])
                self.counts = iter([3, 3, 3, 3, 3, 3, 3])

            def evaluate(self, script):
                if "const labels" in script:
                    return next(self.load_actions)
                if "const scrollable" in script:
                    return next(self.scroll_actions)
                return next(self.counts)

            def wait_for_timeout(self, milliseconds):
                pass

        actions = PageFetcher(max_dynamic_clicks=5, max_dynamic_scrolls=5)._expand_directory(FakePage())
        self.assertEqual(actions, ("load_more",))

    def test_directory_pagination_uses_workflow_fetcher_and_records_each_page(self) -> None:
        class PagingFetcher:
            def __init__(self, root):
                self.root = root
                self.urls = []

            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                self.urls.append((url, expand_directory))
                snapshot = self.root / f"{len(self.urls)}.gz"
                snapshot.write_bytes(b"snapshot")
                html = (
                    "<main><nav><a href='?page=2' rel='next'>Next</a></nav></main>"
                    if url.endswith("/faculty") else "<main>last page</main>"
                )
                return FetchedPage(url, url, 200, "Directory", html, html_to_text(html), str(len(self.urls)), snapshot)

        class PagingCrawler:
            title_pending_records = []

            def __init__(self, timeout):
                pass

            def parse_fetched_directory(self, url, html, status):
                slug = "ada" if url.endswith("/faculty") else "grace"
                name = "Ada Lovelace" if slug == "ada" else "Grace Hopper"
                return [FacultyRecord(name, "Professor", f"https://example.edu/people/{slug}")]

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            policy = DisciplinePolicy("Physics", ("physics",), ())
            task_id = database.create_task(
                policy,
                [SchoolInput("Example University", official_domain="example.edu", directory_url="https://example.edu/faculty")],
                output_dir=root / "output",
                policy_confirmed=True,
            )
            school = database.list_schools(task_id)[0]
            source_id = database.add_source(task_id, school["id"], "https://example.edu/faculty", "faculty_directory", official=True)
            service = WorkflowService(
                database,
                provider=FakeProvider(),
                fetcher=PagingFetcher(root),
                crawler_factory=lambda timeout: PagingCrawler(timeout),
            )
            seeds = {}
            service._collect_directory_seeds(
                task_id, school["id"], database.list_sources(task_id, school["id"])[0], root / "snapshots", seeds, "example.edu"
            )

            self.assertEqual(set(seeds), {"https://example.edu/people/ada", "https://example.edu/people/grace"})
            self.assertEqual(len(database.list_sources(task_id, school["id"])), 2)
            self.assertEqual(source_id, database.list_sources(task_id, school["id"])[0]["id"])

    def test_persisted_successful_directory_snapshot_is_reused(self) -> None:
        class NeverFetcher:
            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                raise AssertionError("persisted successful source should be reused")

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            html = "<article><h3>Ada Lovelace</h3><p>Professor</p><a href='/ada'>Profile</a></article>"
            snapshot = root / "directory.html.gz"
            with gzip.open(snapshot, "wb") as handle:
                handle.write(html.encode("utf-8"))
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput("Example University", official_domain="example.edu")],
                output_dir=root / "output",
                policy_confirmed=True,
            )
            school_id = int(database.list_schools(task_id)[0]["id"])
            source_id = database.add_source(
                task_id, school_id, "https://example.edu/faculty", "faculty_directory",
                official=True, official_boundary="official",
            )
            database.update_source(
                source_id,
                final_url="https://example.edu/faculty",
                http_status=200,
                content_hash="saved-hash",
                snapshot_path=str(snapshot),
                fetch_state="fetched",
            )
            service = WorkflowService(
                database,
                provider=FakeProvider(),
                fetcher=NeverFetcher(),
                crawler_factory=lambda timeout: FakeCrawler(timeout),
            )
            seeds = {}

            service._collect_directory_seeds(
                task_id,
                school_id,
                database.list_sources(task_id, school_id)[0],
                root / "snapshots",
                seeds,
                "example.edu",
            )

            self.assertIn("https://example.edu/ada", seeds)

    def test_linked_official_research_source_persists_graph_metadata(self) -> None:
        class LinkedFetcher(FakeFetcher):
            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                result = super().fetch(url, snapshot_dir, expand_directory=expand_directory)
                if url.endswith("/faculty"):
                    html = "<a href='/research/quantum-lab'>Quantum Research Center</a>"
                elif url.endswith("/research/quantum-lab"):
                    html = "<a href='/people/ada'>Ada Lovelace</a>"
                else:
                    html = "Ada Lovelace Professor of Physics Department of Physics ada@example.edu"
                return FetchedPage(
                    url, url, 200, "Official", html, html_to_text(html), url, result.snapshot_path
                )

        class LinkedCrawler(FakeCrawler):
            def parse_fetched_directory(self, url, html, fetch_status):
                if url.endswith("/research/quantum-lab"):
                    return [FacultyRecord(
                        "Ada Lovelace", "Professor of Physics", "https://example.edu/people/ada",
                        "ada@example.edu",
                    )]
                return []

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput(
                    "Example University",
                    official_domain="example.edu",
                    directory_url="https://example.edu/faculty",
                )],
                output_dir=root / "output",
                policy_confirmed=True,
            )

            WorkflowService(
                database,
                provider=FakeProvider(),
                fetcher=LinkedFetcher(root),
                crawler_factory=lambda timeout: LinkedCrawler(timeout),
            ).run_task(task_id)

            source = next(
                row for row in database.list_sources(task_id)
                if row["source_type"] == "research_unit"
            )
            self.assertEqual(source["discovered_from"], "https://example.edu/faculty")
            self.assertEqual(source["depth"], 1)
            self.assertEqual(source["official_boundary"], "official")
            self.assertEqual(source["fetch_state"], "fetched")

    def test_profile_linked_research_unit_merges_email_by_exact_name(self) -> None:
        class ProfileLinkedFetcher(FakeFetcher):
            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                result = super().fetch(url, snapshot_dir, expand_directory=expand_directory)
                if url.endswith("/faculty"):
                    html = "Ada Lovelace Professor <a href='/people/ada'>Profile</a>"
                elif url.endswith("/people/ada"):
                    html = "Ada Lovelace Professor <a href='/research/quantum-lab'>Quantum Laboratory</a>"
                else:
                    html = "Ada Lovelace Professor <a href='mailto:ada@example.edu'>ada@example.edu</a>"
                return FetchedPage(url, url, 200, "Official", html, html_to_text(html), url, result.snapshot_path)

        class ProfileLinkedCrawler(FakeCrawler):
            def parse_fetched_directory(self, url, html, fetch_status):
                if url.endswith("/faculty"):
                    return [FacultyRecord("Ada Lovelace", "Professor", "https://example.edu/people/ada", "")]
                if url.endswith("/research/quantum-lab"):
                    return [FacultyRecord("Ada Lovelace", "Professor", "", "ada@example.edu")]
                return []

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput("Example University", official_domain="example.edu", directory_url="https://example.edu/faculty")],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )

            summary = WorkflowService(
                database,
                provider=DeepSeekProvider(api_key=""),
                fetcher=ProfileLinkedFetcher(root),
                crawler_factory=lambda timeout: ProfileLinkedCrawler(timeout),
            ).run_task(task_id)

            self.assertEqual(summary["candidates"], {"accepted": 1})
            self.assertEqual(database.list_candidates(task_id, ["accepted"])[0]["email"], "ada@example.edu")
            self.assertIn("research_unit", {row["source_type"] for row in database.list_sources(task_id)})

    def test_search_discovery_refetches_only_official_candidate_url(self) -> None:
        class UrlOnlyDiscovery:
            def discover(self, *, name, school, official_domain):
                return (
                    DiscoveryHint("https://outside.test/ada", f"{name} {school}"),
                    DiscoveryHint("https://research.example.edu/people/ada", f"{name} {school}"),
                )

        class DirectoryWithoutProfileCrawler(FakeCrawler):
            def parse_fetched_directory(self, url, html, fetch_status):
                return [FacultyRecord("Ada Lovelace", "Professor of Physics", "", "")]

        class TrackingFetcher(FakeFetcher):
            def __init__(self, root):
                super().__init__(root)
                self.urls = []

            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                self.urls.append(url)
                result = super().fetch(url, snapshot_dir, expand_directory=expand_directory)
                if url == "https://research.example.edu/people/ada":
                    html = (
                        "Ada Lovelace Professor of Physics Department of Physics "
                        "<a href='mailto:ada@research.example.edu'>ada@research.example.edu</a>"
                    )
                    return FetchedPage(
                        url, url, 200, "Ada Lovelace", html, html_to_text(html),
                        "official-profile", result.snapshot_path,
                    )
                return result

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            database = WorkflowDatabase(root / "workflow.db")
            task_id = database.create_task(
                DisciplinePolicy("Physics", ("physics",), ()),
                [SchoolInput(
                    "Example University",
                    official_domain="example.edu",
                    directory_url="https://example.edu/faculty",
                )],
                output_dir=root / "output",
                routine_model="local-only",
                escalation_model="local-only",
                policy_confirmed=True,
            )
            fetcher = TrackingFetcher(root)

            summary = WorkflowService(
                database,
                provider=DeepSeekProvider(api_key=""),
                fetcher=fetcher,
                crawler_factory=lambda timeout: DirectoryWithoutProfileCrawler(timeout),
                discovery_provider=UrlOnlyDiscovery(),
            ).run_task(task_id)

            self.assertNotIn("https://outside.test/ada", fetcher.urls)
            self.assertIn("https://research.example.edu/people/ada", fetcher.urls)
            self.assertEqual(summary["candidates"], {"accepted": 1})
            candidate = database.list_candidates(task_id, ["accepted"])[0]
            self.assertEqual(candidate["email"], "ada@research.example.edu")
            evidence = json.loads(candidate["evidence_json"])
            self.assertTrue(all("example.edu" in item["source_url"] for item in evidence))
            self.assertTrue(any(item["source_url"].endswith("/people/ada") for item in evidence))
            self.assertFalse(any("outside.test" in item["source_url"] for item in evidence))

    def test_access_block_is_saved_as_a_redacted_source_diagnostic(self) -> None:
        class BlockedFetcher:
            def fetch(self, url, snapshot_dir, *, expand_directory=False):
                raise AccessBlockedError("Page requires human verification")

        with tempfile.TemporaryDirectory() as temp_dir:
            database = WorkflowDatabase(Path(temp_dir) / "workflow.db")
            policy = DisciplinePolicy("Physics", ("physics",), ())
            task_id = database.create_task(
                policy,
                [SchoolInput("Example University", official_domain="example.edu", directory_url="https://example.edu/faculty")],
                output_dir=temp_dir,
                policy_confirmed=True,
            )
            summary = WorkflowService(database, provider=FakeProvider(), fetcher=BlockedFetcher()).run_task(task_id)

            self.assertEqual(summary["schools"], {"blocked_access": 1})
            self.assertEqual(
                database.list_sources(task_id)[0]["failure_reason"],
                "Page requires human verification",
            )
            review = database.list_access_reviews(task_id)[0]
            self.assertEqual(review["url"], "https://example.edu/faculty")
            self.assertEqual(review["status"], "pending")


if __name__ == "__main__":
    unittest.main()
